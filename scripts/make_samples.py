"""Generate a deliberately messy demo corpus exercising every pipeline path:

- sales_report.xlsx    : merged title, two-row merged header, side-by-side +
                         stacked tables on one sheet, second sheet with notes
- transactions.csv     : cp1252, semicolon delimiter, Swiss thousands
                         separators, one ragged row
- transactions_april.csv: same schema again (schema-memory hit + drift check)
- customers.csv        : latin-1, joins to transactions on customer id
- returns_q2.csv       : same schema as the xlsx returns table (cross-format
                         schema-memory hit)
- invoice.pdf          : page 1 native text + ruled table; page 2 is the same
                         content embedded as an image (scanned simulation)
"""

from pathlib import Path

import openpyxl
import pymupdf
from openpyxl.styles import Font

OUT = Path(__file__).resolve().parent.parent / "samples"
OUT.mkdir(exist_ok=True)

REGIONS = ["Zürich", "Bern", "Basel", "Genf", "Luzern"]


def make_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"

    ws.merge_cells("A1:F1")
    ws["A1"] = "ACME GmbH — Regional Sales Report H1 2025 (CHF)"
    ws["A1"].font = Font(bold=True, size=14)

    # Table A (B3:F9): two-row header, merged spans.
    ws.merge_cells("B3:B4")
    ws["B3"] = "Region"
    ws.merge_cells("C3:D3")
    ws["C3"] = "Q1 2025"
    ws.merge_cells("E3:F3")
    ws["E3"] = "Q2 2025"
    for col, label in zip("CDEF", ["Units", "Revenue", "Units", "Revenue"], strict=True):
        ws[f"{col}4"] = label
    data_a = [
        ("Zürich", 1240, 812400, 1315, 861200),
        ("Bern", 720, 455800, 698, 442100),
        ("Basel", 810, 529300, 934, 601000),
        ("Genf", 505, 358900, 541, 371400),
        ("Luzern", 330, 201500, 361, 224800),
    ]
    for i, row in enumerate(data_a, start=5):
        for j, v in enumerate(row):
            ws.cell(row=i, column=2 + j, value=v)

    # Table B (H3:J9): side-by-side product list, one blank column away.
    for j, h in enumerate(["Product", "Category", "Unit Price"]):
        ws.cell(row=3, column=8 + j, value=h)
    products = [("TwinFlow 200", "Pumps", 1250.0), ("TwinFlow 300", "Pumps", 1690.0),
                ("SenseKit v2", "Sensors", 420.0), ("SenseKit Pro", "Sensors", 780.0),
                ("FlexMount", "Accessories", 95.5), ("BaseRail 1m", "Accessories", 45.0)]
    for i, row in enumerate(products, start=4):
        for j, v in enumerate(row):
            ws.cell(row=i, column=8 + j, value=v)

    # Table C (B13:D18): stacked below A after a gap — returns by region.
    for j, h in enumerate(["Region", "Returns", "Return Rate"]):
        ws.cell(row=13, column=2 + j, value=h)
    returns = [("Zürich", 41, 0.032), ("Bern", 12, 0.017), ("Basel", 25, 0.029),
               ("Genf", 9, 0.017), ("Luzern", 6, 0.017)]
    for i, row in enumerate(returns, start=14):
        for j, v in enumerate(row):
            ws.cell(row=i, column=2 + j, value=v)

    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Data notes"
    ws2["A3"] = "Field"
    ws2["B3"] = "Description"
    notes = [("Revenue", "Gross revenue in CHF, excl. VAT"),
             ("Returns", "Units returned within 30 days"),
             ("Region", "Swiss sales region (canton hub)")]
    for i, (a, b) in enumerate(notes, start=4):
        ws2[f"A{i}"] = a
        ws2[f"B{i}"] = b

    wb.save(OUT / "sales_report.xlsx")


def make_csvs():
    header = "TxnID;CustomerID;Datum;Betrag CHF;Region;Beschreibung"
    rows_march = [
        "T-1001;CUST-001;2025-03-03;1'250.50;Zürich;Wartung Pumpenanlage",
        "T-1002;CUST-002;2025-03-05;890.00;Bern;Sensorik Nachrüstung",
        "T-1003;CUST-001;2025-03-09;2'340.00;Zürich;TwinFlow 300 Installation",
        "T-1004;CUST-003;2025-03-11;420.00;Basel;SenseKit v2",
        "T-1005;CUST-004;2025-03-14;779.90;Genf;Réparation façade capteur",
        # ragged row (extra field sneaks in):
        "T-1006;CUST-002;2025-03-18;1'690.00;Bern;Ersatzteil;EXPRESS",
        "T-1007;CUST-005;2025-03-21;95.50;Luzern;FlexMount Halterung",
        "T-1008;CUST-003;2025-03-27;5'120.75;Basel;Grossauftrag Ventile",
    ]
    (OUT / "transactions.csv").write_bytes(
        ("\n".join([header] + rows_march) + "\n").encode("cp1252"))

    rows_april = [
        "T-1101;CUST-001;2025-04-02;980.00;Zürich;Servicevertrag Q2",
        "T-1102;CUST-004;2025-04-08;1'420.00;Genf;TwinFlow 200",
        "T-1103;CUST-005;2025-04-15;240.00;Luzern;BaseRail Montage",
        "T-1104;CUST-002;2025-04-19;3'310.00;Bern;Anlagenerweiterung",
        "T-1105;CUST-003;2025-04-25;660.40;Basel;SenseKit Pro Upgrade",
    ]
    (OUT / "transactions_april.csv").write_bytes(
        ("\n".join([header] + rows_april) + "\n").encode("cp1252"))

    customers = [
        "customer_id,name,region,segment",
        "CUST-001,Müller AG,Zürich,Industrie",
        "CUST-002,Bärtschi & Söhne,Bern,KMU",
        "CUST-003,Rheinwerk GmbH,Basel,Industrie",
        "CUST-004,Café Rivière SA,Genf,Gastro",
        "CUST-005,Alpenblick Technik,Luzern,KMU",
    ]
    (OUT / "customers.csv").write_bytes(("\n".join(customers) + "\n").encode("latin-1"))

    returns_q2 = [
        "Region,Returns,Return Rate",
        "Zürich,38,0.029",
        "Bern,15,0.021",
        "Basel,22,0.024",
        "Genf,11,0.020",
        "Luzern,8,0.022",
    ]
    (OUT / "returns_q2.csv").write_bytes(("\n".join(returns_q2) + "\n").encode("utf-8"))


def make_pdf():
    doc = pymupdf.open()
    page = doc.new_page()  # A4-ish default 595x842
    x0 = 72

    page.insert_text((x0, 80), "INVOICE  INV-2025-0042", fontsize=16)
    page.insert_text((x0, 110), "ACME GmbH · Technoparkstrasse 1 · 8005 Zürich", fontsize=9)
    page.insert_text((x0, 140), "Bill to: Müller AG, Bahnhofstrasse 10, 8001 Zürich", fontsize=10)
    page.insert_text((x0, 155), "Customer ID: CUST-001 · Date: 2025-03-31 · Terms: 30 days net",
                     fontsize=10)

    # Ruled table so find_tables() has explicit lines.
    cols = [72, 300, 360, 450, 520]
    rows_y = [200, 220, 240, 260, 280, 300]
    for y in rows_y:
        page.draw_line((cols[0], y), (cols[-1], y))
    for x in cols:
        page.draw_line((x, rows_y[0]), (x, rows_y[-1]))
    header = ["Item", "Qty", "Unit Price", "Total"]
    body = [
        ["TwinFlow 300 pump", "1", "1'690.00", "1'690.00"],
        ["SenseKit v2", "2", "420.00", "840.00"],
        ["Installation (hours)", "6", "140.00", "840.00"],
        ["", "", "Total CHF", "3'370.00"],
    ]
    def put_row(vals, y):
        for i, v in enumerate(vals):
            page.insert_text((cols[i] + 4, y + 14), v, fontsize=9)
    put_row(header, rows_y[0])
    for i, r in enumerate(body):
        put_row(r, rows_y[i + 1])

    page.insert_text((x0, 340),
                     "Payment within 30 days to IBAN CH93 0076 2011 6238 5295 7.", fontsize=9)
    page.insert_text((x0, 355),
                     "All amounts in CHF excl. 8.1% VAT. Warranty per SIA 118 terms.", fontsize=9)

    # Page 2: the same page as a full-page image -> classifier must call it scanned.
    pix = page.get_pixmap(dpi=150)
    page2 = doc.new_page(width=page.rect.width, height=page.rect.height)
    page2.insert_image(page2.rect, pixmap=pix)

    doc.save(OUT / "invoice.pdf", deflate=True, deflate_images=True, garbage=3)
    doc.close()


if __name__ == "__main__":
    make_xlsx()
    make_csvs()
    make_pdf()
    print(f"samples written to {OUT}")
    for p in sorted(OUT.iterdir()):
        print(" -", p.name, p.stat().st_size, "bytes")
