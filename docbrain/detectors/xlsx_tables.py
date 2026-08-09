"""XLSX track, stage 1: heuristic table detection ("island detection").

Same problem class eparse/TableSense address: multiple tables per sheet,
irregular offsets, merged headers. Approach: build a boolean occupancy grid
(propagating merged-cell values so headers stay connected), find 8-connected
components, merge overlapping bounding boxes, then infer header rows per island.
Islands that look ambiguous get flagged so the agent loop can refine them."""

from __future__ import annotations

from collections import deque
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class Island:
    sheet: str
    r0: int  # 0-based inclusive
    c0: int
    r1: int  # inclusive
    c1: int
    grid: list[list]                       # raw values incl. header rows
    header_rows: int = 1
    merged_in_header: bool = False
    flags: list[str] = field(default_factory=list)

    @property
    def ref(self) -> str:
        return (f"{self.sheet}!{get_column_letter(self.c0 + 1)}{self.r0 + 1}:"
                f"{get_column_letter(self.c1 + 1)}{self.r1 + 1}")

    @property
    def n_rows(self) -> int:
        return self.r1 - self.r0 + 1

    @property
    def n_cols(self) -> int:
        return self.c1 - self.c0 + 1


def _is_filled(v) -> bool:
    return v is not None and str(v).strip() != ""


def _load_grid(ws) -> tuple[list[list], set[tuple[int, int, int, int]]]:
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    merged: set[tuple[int, int, int, int]] = set()
    for rng in ws.merged_cells.ranges:
        r0, c0 = rng.min_row - 1, rng.min_col - 1
        r1, c1 = rng.max_row - 1, rng.max_col - 1
        merged.add((r0, c0, r1, c1))
        if r0 < len(grid) and c0 < len(grid[0] if grid else []):
            top_left = grid[r0][c0]
            for r in range(r0, min(r1 + 1, len(grid))):
                for c in range(c0, min(c1 + 1, len(grid[r]))):
                    grid[r][c] = top_left
    return grid, merged


def _components(grid: list[list]) -> list[tuple[int, int, int, int]]:
    if not grid:
        return []
    R, C = len(grid), max(len(r) for r in grid)
    filled = [[_is_filled(grid[r][c]) if c < len(grid[r]) else False for c in range(C)]
              for r in range(R)]
    seen = [[False] * C for _ in range(R)]
    boxes = []
    for r in range(R):
        for c in range(C):
            if filled[r][c] and not seen[r][c]:
                q = deque([(r, c)])
                seen[r][c] = True
                rr0 = rr1 = r
                cc0 = cc1 = c
                while q:
                    y, x = q.popleft()
                    rr0, rr1 = min(rr0, y), max(rr1, y)
                    cc0, cc1 = min(cc0, x), max(cc1, x)
                    for dy in (-1, 0, 1):
                        for dx in (-1, 0, 1):
                            ny, nx = y + dy, x + dx
                            if 0 <= ny < R and 0 <= nx < C and filled[ny][nx] and not seen[ny][nx]:
                                seen[ny][nx] = True
                                q.append((ny, nx))
                boxes.append((rr0, cc0, rr1, cc1))
    return _merge_overlaps(boxes)


def _merge_overlaps(boxes: list[tuple[int, int, int, int]]) -> list[tuple[int, int, int, int]]:
    changed = True
    boxes = list(boxes)
    while changed:
        changed = False
        out: list[tuple[int, int, int, int]] = []
        while boxes:
            b = boxes.pop()
            merged_any = False
            for i, o in enumerate(out):
                if not (b[2] < o[0] or o[2] < b[0] or b[3] < o[1] or o[3] < b[1]):
                    out[i] = (min(b[0], o[0]), min(b[1], o[1]), max(b[2], o[2]), max(b[3], o[3]))
                    merged_any = changed = True
                    break
            if not merged_any:
                out.append(b)
        boxes = out if changed else []
        if not changed:
            return out
    return boxes


def _infer_header_rows(sub: list[list]) -> tuple[int, list[str]]:
    """How many leading rows are headers? Returns (n_header_rows, flags)."""
    flags: list[str] = []

    def str_ratio(row) -> float:
        vals = [v for v in row if _is_filled(v)]
        if not vals:
            return 0.0
        return sum(isinstance(v, str) for v in vals) / len(vals)

    def numeric_ratio(row) -> float:
        vals = [v for v in row if _is_filled(v)]
        if not vals:
            return 0.0
        return sum(isinstance(v, (int, float)) and not isinstance(v, bool) for v in vals) / len(vals)

    def has_horizontal_spans(row) -> bool:
        """Adjacent duplicate values — the signature of merge-filled header spans."""
        vals = [str(v).strip() for v in row if _is_filled(v)]
        return any(a == b for a, b in zip(vals, vals[1:], strict=False))

    if len(sub) < 2:
        return 0, ["single_row_island"]

    if not (str_ratio(sub[0]) >= 0.6 and numeric_ratio(sub[0]) <= 0.2):
        flags.append("no_obvious_header")
        return (1 if str_ratio(sub[0]) > numeric_ratio(sub[0]) else 0), flags

    n = 1
    # Extend to a 2nd/3rd header row only with positive evidence: the extra row
    # is string-typed, a spanning pattern exists, and a typed body sits below.
    for i in range(1, min(3, len(sub) - 1)):
        body = sub[i + 1:]
        body_numeric = (sum(numeric_ratio(r) for r in body) / len(body)) if body else 0.0
        if (str_ratio(sub[i]) >= 0.8
                and (has_horizontal_spans(sub[i - 1]) or has_horizontal_spans(sub[i]))
                and body_numeric >= 0.2):
            n = i + 1
        else:
            break
    if n >= 2:
        flags.append("multi_row_header")
    return n, flags


def detect_islands(path, min_cells: int = 4) -> list[Island]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    islands: list[Island] = []
    for ws in wb.worksheets:
        grid, merged = _load_grid(ws)
        if not grid:
            continue
        for (r0, c0, r1, c1) in _components(grid):
            n_rows, n_cols = r1 - r0 + 1, c1 - c0 + 1
            if n_rows * n_cols < min_cells or n_cols < 2:
                continue  # scrap: titles, footnotes, lone cells
            sub = [[grid[r][c] if c < len(grid[r]) else None for c in range(c0, c1 + 1)]
                   for r in range(r0, r1 + 1)]
            header_rows, flags = _infer_header_rows(sub)
            merged_in_header = any(
                mr0 >= r0 and mr1 <= r0 + max(header_rows, 1) - 1 and mc0 >= c0 and mc1 <= c1
                for (mr0, mc0, mr1, mc1) in merged)
            if merged_in_header and header_rows < 2:
                flags.append("merged_header_uncertain")
            isl = Island(sheet=ws.title, r0=r0, c0=c0, r1=r1, c1=c1, grid=sub,
                         header_rows=header_rows, merged_in_header=merged_in_header,
                         flags=flags)
            islands.append(isl)
    wb.close()
    return islands
